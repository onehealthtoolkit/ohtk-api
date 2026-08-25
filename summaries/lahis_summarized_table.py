"""
LAHIS-only summarized dashboard table (FAO template + appended census context).

The original A–AN grid remains stable; household, village-census, and close
attribution columns are appended at AO–AW.
Not used by generic OHTK tenants.
"""

from __future__ import annotations

from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Fixed report type for this LAHIS export (not a download parameter).
DEFAULT_REPORT_TYPE_NAME = "Animal Sick/Death"

# Template species headers (row 3), left-to-right within each metric block.
SPECIES_HEADERS: Tuple[str, ...] = (
    "Buffalo",
    "Cattle",
    "pig",
    "Goat-Sheep",
    "Chicken",
    "Duck",
)
VILLAGE_SPECIES_HEADERS: Tuple[str, ...] = (
    "Buffalo",
    "Cattle",
    "pig",
    "Goat-Sheep",
    "Chicken",
    "Duck / other poultry",
)

# Group labels (row 2) — fixed wording per product locks.
GROUP_VILLAGE = "Village information"
GROUP_DATE = "Date and test result"
GROUP_AFFECTED_POPULATION = "Animals in affected households"
GROUP_SICK = "Animal sick"
GROUP_DEAD = "Animal dead"
GROUP_RECOVERD = "Animal recoverd"
GROUP_STAMPED = "Stamped out"
GROUP_HOUSEHOLDS = "Households"
GROUP_VILLAGE_POPULATION = "Village animal population"
GROUP_CASE_CLOSURE = "Case closure"

# Column indexes 1-based for openpyxl.
COL_ID = 1
COL_PROVINCE = 2
COL_DISTRICT = 3
COL_VILLAGE = 4
COL_LAT = 5
COL_LNG = 6
COL_STARTED = 7
COL_STOPPED = 8
COL_SUSPECTED = 9
COL_TEST_RESULT = 10
COL_POP_START = 11  # K
COL_SICK_START = 17  # Q
COL_DEAD_START = 23  # W
COL_RECOVER_START = 29  # AC
COL_STAMP_START = 35  # AI
COL_AFFECTED_HOUSEHOLDS = 41  # AO
COL_VILLAGE_HOUSEHOLDS = 42  # AP
COL_VILLAGE_POP_START = 43  # AQ
COL_CLOSE_SOURCE = 49  # AW
TOTAL_COLS = 49  # A..AW


def normalize_species(raw: Any) -> Optional[str]:
    """Map report animal_species to a template column header, or None if unmapped."""
    if raw is None:
        return None
    s = str(raw).strip().lower().replace("_", "-").replace(" ", "")
    if not s:
        return None
    # multi-select leftovers: take first token
    if "," in s:
        s = s.split(",")[0].strip()
    mapping = {
        "buffalo": "Buffalo",
        "cattle": "Cattle",
        "pig": "pig",
        "goat": "Goat-Sheep",
        "sheep": "Goat-Sheep",
        "goat-sheep": "Goat-Sheep",
        "goatsheep": "Goat-Sheep",
        "chicken": "Chicken",
        "duck": "Duck",
    }
    return mapping.get(s)


def species_column_offset(species_header: str) -> Optional[int]:
    try:
        return SPECIES_HEADERS.index(species_header)
    except ValueError:
        return None


def _as_number(value: Any) -> Any:
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        return value
    try:
        s = str(value).strip()
        if s == "":
            return ""
        if "." in s:
            return float(s)
        return int(s)
    except (TypeError, ValueError):
        return value


def _format_dt(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        try:
            local = value.astimezone() if value.tzinfo else value
            return local.strftime("%d-%b-%Y %H:%M:%S")
        except Exception:
            return value.strftime("%d-%b-%Y %H:%M:%S")
    return str(value)


def format_close_source(value: Any) -> str:
    """Return a human-readable lifecycle close source for Excel."""
    source = str(value or "").strip().lower()
    return {"officer": "Officer", "system": "System"}.get(source, "")


def _data_dict(data: Any) -> dict:
    return data if isinstance(data, dict) else {}


def metric_from_data_or_accum(
    data: dict, accumulated: Any, field: str
) -> Any:
    """Prefer accumulated metric total when present; else form data field."""
    if isinstance(accumulated, dict):
        metrics = accumulated.get("metrics") or []
        if isinstance(metrics, list):
            for m in metrics:
                if not isinstance(m, dict):
                    continue
                mid = m.get("id") or m.get("reportField")
                if mid == field and m.get("value") is not None:
                    return _as_number(m.get("value"))
    return _as_number(data.get(field))


def parse_gps(gps_location_str: Optional[str]) -> Tuple[str, str]:
    """
    GeoDjango Point string is 'x,y' = longitude,latitude.
    Template wants Latitude then Longitude.
    """
    if not gps_location_str:
        return "", ""
    parts = str(gps_location_str).split(",")
    if len(parts) < 2:
        return "", ""
    lng = parts[0].strip()
    lat = parts[1].strip()
    return lat, lng


def resolve_province_district(authority) -> Tuple[str, str]:
    """
    From leaf relevant authority: district = leaf, province = first parent.
    Village is not an authority layer in LAHIS seeds — left to caller.
    """
    if authority is None:
        return "", ""
    parents = list(authority.inherits.all()) if hasattr(authority, "inherits") else []
    if not parents:
        return authority.name or "", ""
    parent = parents[0]
    # If parent itself has a parent, treat parent as province and leaf as district.
    return (parent.name or ""), (authority.name or "")


def resolve_report_village_name(
    report, fallback_by_user: Dict[Any, str]
) -> str:
    """Prefer the report's selected village; fall back for legacy reports."""
    village = getattr(report, "village", None)
    selected_name = getattr(village, "name", "") if village is not None else ""
    if selected_name:
        return selected_name
    return fallback_by_user.get(getattr(report, "reported_by_id", None), "")


def resolve_report_census_snapshot(report, snapshots_by_village: Dict[Any, list]):
    """Return the newest village census on or before the incident date."""
    village_id = getattr(report, "village_id", None)
    if not village_id:
        return None
    snapshots = snapshots_by_village.get(village_id, [])
    incident_date = getattr(report, "incident_date", None)
    if isinstance(incident_date, datetime):
        incident_date = incident_date.date()
    if incident_date is None:
        return snapshots[0] if snapshots else None
    for snapshot in snapshots:
        if snapshot.census_date <= incident_date:
            return snapshot
    return None


def census_values_for_snapshot(snapshot) -> Tuple[Any, Dict[str, Any]]:
    """Extract village households and template-species population."""
    if snapshot is None:
        return "", {}

    form_data = snapshot.form_data if isinstance(snapshot.form_data, dict) else {}
    summary = form_data.get("summary")
    summary = summary if isinstance(summary, dict) else {}
    village_households = _as_number(summary.get("village_household_quantity"))

    population: Dict[str, Any] = {}
    species_mapping = {
        "BUFFALO": "Buffalo",
        "CATTLE": "Cattle",
        "PIG": "pig",
        "GOAT": "Goat-Sheep",
        "SHEEP": "Goat-Sheep",
        "CHICKEN": "Chicken",
        "OTHER_POULTRY": "Duck / other poultry",
        "DUCK": "Duck / other poultry",
    }
    for fact in snapshot.facts.all():
        measures = fact.measures if isinstance(fact.measures, dict) else {}
        value = _as_number(measures.get("animal_quantity"))
        if not isinstance(value, (int, float)) or isinstance(value, bool):
            continue
        dimensions = (
            fact.extra_dimensions
            if isinstance(fact.extra_dimensions, dict)
            else {}
        )
        species_key = dimensions.get("species")
        if not species_key and str(fact.row_key).startswith("species:"):
            species_key = str(fact.row_key).split(":", 1)[1]
        header = species_mapping.get(str(species_key or "").upper())
        if not header:
            continue
        existing = population.get(header, 0)
        population[header] = existing + value

    return village_households, population


def build_row_values(
    *,
    report_id: Any,
    data: Any,
    incident_date: Any,
    stopped_at: Any,
    ai_suspected: Any,
    test_result: Any,
    stamp_out: Any,
    gps_location_str: Optional[str],
    province: str,
    district: str,
    village: str,
    accumulated_metrics: Any = None,
    village_households: Any = None,
    village_animal_population: Optional[Dict[str, Any]] = None,
    close_source: Any = None,
) -> List[Any]:
    """Return 49 cell values (A..AW) for one report row."""
    data = _data_dict(data)
    species_raw = data.get("animal_species")
    species = normalize_species(species_raw)
    lat, lng = parse_gps(gps_location_str)

    population = metric_from_data_or_accum(
        data, accumulated_metrics, "num_total_animal"
    )
    sick = metric_from_data_or_accum(data, accumulated_metrics, "num_sick")
    dead = metric_from_data_or_accum(data, accumulated_metrics, "num_dead")
    recover = metric_from_data_or_accum(data, accumulated_metrics, "num_recover")
    affected_households = metric_from_data_or_accum(
        data, accumulated_metrics, "num_household"
    )
    stamp = _as_number(stamp_out if stamp_out is not None else data.get("stamp_out"))

    row: List[Any] = [""] * TOTAL_COLS
    row[COL_ID - 1] = str(report_id) if report_id is not None else ""
    row[COL_PROVINCE - 1] = province or ""
    row[COL_DISTRICT - 1] = district or ""
    row[COL_VILLAGE - 1] = village or ""
    row[COL_LAT - 1] = lat
    row[COL_LNG - 1] = lng
    row[COL_STARTED - 1] = _format_dt(incident_date)
    row[COL_STOPPED - 1] = _format_dt(stopped_at)
    row[COL_SUSPECTED - 1] = ai_suspected or ""
    row[COL_TEST_RESULT - 1] = test_result or ""

    if species:
        off = species_column_offset(species)
        if off is not None:
            row[COL_POP_START - 1 + off] = population
            row[COL_SICK_START - 1 + off] = sick
            row[COL_DEAD_START - 1 + off] = dead
            row[COL_RECOVER_START - 1 + off] = recover
            row[COL_STAMP_START - 1 + off] = stamp

    row[COL_AFFECTED_HOUSEHOLDS - 1] = affected_households
    row[COL_VILLAGE_HOUSEHOLDS - 1] = _as_number(village_households)
    for species_header, value in (village_animal_population or {}).items():
        try:
            off = VILLAGE_SPECIES_HEADERS.index(species_header)
        except ValueError:
            continue
        row[COL_VILLAGE_POP_START - 1 + off] = _as_number(value)
    row[COL_CLOSE_SOURCE - 1] = format_close_source(close_source)

    return row


def write_header_rows(ws) -> None:
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Row 2 groups
    groups = [
        (COL_PROVINCE, COL_LNG, GROUP_VILLAGE),
        (COL_STARTED, COL_TEST_RESULT, GROUP_DATE),
        (COL_POP_START, COL_POP_START + 5, GROUP_AFFECTED_POPULATION),
        (COL_SICK_START, COL_SICK_START + 5, GROUP_SICK),
        (COL_DEAD_START, COL_DEAD_START + 5, GROUP_DEAD),
        (COL_RECOVER_START, COL_RECOVER_START + 5, GROUP_RECOVERD),
        (COL_STAMP_START, COL_STAMP_START + 5, GROUP_STAMPED),
        (
            COL_AFFECTED_HOUSEHOLDS,
            COL_VILLAGE_HOUSEHOLDS,
            GROUP_HOUSEHOLDS,
        ),
        (
            COL_VILLAGE_POP_START,
            COL_VILLAGE_POP_START + 5,
            GROUP_VILLAGE_POPULATION,
        ),
        (COL_CLOSE_SOURCE, COL_CLOSE_SOURCE, GROUP_CASE_CLOSURE),
    ]
    for start, end, title in groups:
        if start == end:
            cell = ws.cell(2, start, title)
        else:
            ws.merge_cells(
                start_row=2,
                start_column=start,
                end_row=2,
                end_column=end,
            )
            cell = ws.cell(2, start, title)
        cell.font = bold
        cell.alignment = center

    # Row 3 column headers
    headers = [
        "ID",
        "Province",
        "District",
        "Village",
        "Latitude",
        "Longitude",
        "started date",
        "stopped date",
        "suspected",
        "Test result",
    ]
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(3, idx, h)
        cell.font = bold
        cell.alignment = center

    for block_start in (
        COL_POP_START,
        COL_SICK_START,
        COL_DEAD_START,
        COL_RECOVER_START,
        COL_STAMP_START,
    ):
        for i, species in enumerate(SPECIES_HEADERS):
            cell = ws.cell(3, block_start + i, species)
            cell.font = bold
            cell.alignment = center

    for i, species in enumerate(VILLAGE_SPECIES_HEADERS):
        cell = ws.cell(3, COL_VILLAGE_POP_START + i, species)
        cell.font = bold
        cell.alignment = center

    ws.cell(3, COL_AFFECTED_HOUSEHOLDS, "Affected households").font = bold
    ws.cell(3, COL_AFFECTED_HOUSEHOLDS).alignment = center
    ws.cell(3, COL_VILLAGE_HOUSEHOLDS, "Village households").font = bold
    ws.cell(3, COL_VILLAGE_HOUSEHOLDS).alignment = center
    ws.cell(3, COL_CLOSE_SOURCE, "Close source").font = bold
    ws.cell(3, COL_CLOSE_SOURCE).alignment = center


def build_workbook(rows: Sequence[Sequence[Any]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summarized table"
    write_header_rows(ws)
    for r_i, values in enumerate(rows, start=4):
        for c_i, value in enumerate(values, start=1):
            if value is None or value == "":
                continue
            ws.cell(r_i, c_i, value)
    for col in range(1, TOTAL_COLS + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["AW"].width = 14
    return wb


def row_from_incident_report(
    report, case=None, village_name: str = "", census_snapshot=None
) -> List[Any]:
    """Build one export row from IncidentReport (+ optional Case)."""
    from reports.metric_accumulation import accumulate_incident_metrics

    data = report.data if isinstance(report.data, dict) else {}
    authority = None
    try:
        authority = report.relevant_authorities.all().first()
    except Exception:
        authority = None
    province, district = resolve_province_district(authority)

    stopped_at = None
    test_result = ""
    stamp_out = None
    close_source = ""
    if case is not None:
        stopped_at = getattr(case, "stopped_at", None)
        close_source = getattr(case, "close_source", "") or ""
        payload = getattr(case, "close_payload", None) or {}
        if isinstance(payload, dict):
            test_result = payload.get("test_result") or ""
            if payload.get("stamp_out") is not None:
                stamp_out = payload.get("stamp_out")
        # Property projects Layer2 test_result
        if not test_result:
            try:
                test_result = case.test_result or ""
            except Exception:
                pass

    try:
        accumulated = accumulate_incident_metrics(report)
    except Exception:
        accumulated = None
    village_households, village_animal_population = census_values_for_snapshot(
        census_snapshot
    )

    return build_row_values(
        report_id=report.id,
        data=data,
        incident_date=getattr(report, "incident_date", None),
        stopped_at=stopped_at,
        ai_suspected=getattr(report, "ai_suspected", "") or "",
        test_result=test_result,
        stamp_out=stamp_out,
        gps_location_str=getattr(report, "gps_location_str", "") or "",
        province=province,
        district=district,
        village=village_name or "",
        accumulated_metrics=accumulated,
        village_households=village_households,
        village_animal_population=village_animal_population,
        close_source=close_source,
    )


def resolve_fixed_report_type():
    """
    Always Animal Sick/Death for this export.
    Optional tenant Configuration key cases.lahis_summarized_report_type_name.
    """
    from accounts.models import Configuration
    from reports.models.report_type import ReportType

    name = DEFAULT_REPORT_TYPE_NAME
    try:
        configured = Configuration.get("cases.lahis_summarized_report_type_name")
        if configured and str(configured).strip():
            name = str(configured).strip()
    except Exception:
        pass
    return ReportType.objects.get(name=name)


def collect_export_rows(
    *,
    from_date=None,
    to_date=None,
    report_type=None,
) -> List[List[Any]]:
    """
    Query all non-test reports of the fixed LAHIS report type (tenant-wide).
    No authority filter — operators filter in Excel if needed.
    """
    from cases.models import Case
    from reports.models.report import IncidentReport

    if report_type is None:
        report_type = resolve_fixed_report_type()

    qs = (
        IncidentReport.objects.filter(
            test_flag=False,
            report_type=report_type,
        )
        .select_related("report_type", "village")
        .prefetch_related("relevant_authorities", "relevant_authorities__inherits")
        .order_by("-created_at")
        .distinct()
    )
    if from_date:
        qs = qs.filter(created_at__gte=from_date)
    if to_date:
        qs = qs.filter(created_at__lte=to_date)

    reports = list(qs)
    case_ids = [r.case_id for r in reports if r.case_id]
    case_by_id: Dict[str, Any] = {
        str(c.id): c
        for c in Case.objects.filter(id__in=case_ids).only(
            "id",
            "stopped_at",
            "close_payload",
            "close_source",
        )
    }

    snapshots_by_village: Dict[Any, list] = {}
    village_ids = {r.village_id for r in reports if r.village_id}
    if village_ids:
        try:
            from census.models import (
                CensusDefinition,
                CensusRoundDefinition,
                VillageCensusSnapshot,
            )

            snapshots = (
                VillageCensusSnapshot.objects.filter(
                    village_id__in=village_ids,
                    status=VillageCensusSnapshot.Status.SUBMITTED,
                    round_occurrence__kind=CensusDefinition.Kind.ANIMAL,
                    round_occurrence__mode=CensusRoundDefinition.Mode.PRODUCTION,
                )
                .prefetch_related("facts")
                .order_by("village_id", "-census_date", "-submitted_at", "-id")
            )
            for snapshot in snapshots:
                snapshots_by_village.setdefault(snapshot.village_id, []).append(
                    snapshot
                )
        except Exception:
            pass

    # Legacy fallback for reports created before the selected village was stored.
    village_by_user: Dict[Any, str] = {}
    try:
        from accounts.models import VillageReporterAssignment

        user_ids = {r.reported_by_id for r in reports if r.reported_by_id}
        for uv in VillageReporterAssignment.objects.filter(
            reporter_id__in=user_ids
        ).select_related("village"):
            if uv.reporter_id not in village_by_user and uv.village_id:
                village_by_user[uv.reporter_id] = uv.village.name
    except Exception:
        pass

    rows: List[List[Any]] = []
    for report in reports:
        case = case_by_id.get(str(report.case_id)) if report.case_id else None
        village = resolve_report_village_name(report, village_by_user)
        census_snapshot = resolve_report_census_snapshot(
            report, snapshots_by_village
        )
        rows.append(
            row_from_incident_report(
                report,
                case=case,
                village_name=village,
                census_snapshot=census_snapshot,
            )
        )
    return rows
